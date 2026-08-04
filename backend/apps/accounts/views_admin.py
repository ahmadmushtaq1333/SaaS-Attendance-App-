from rest_framework import viewsets, status, views
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.accounts.permissions import IsAdminUser, IsGlobalAdmin
from apps.institutions.models import Institution, Department, Semester, Section
from apps.courses.models import Course, Enrollment
from apps.attendance.models import AttendanceSession
from django.contrib.auth import get_user_model
from django.db.models import Count, Q
from rest_framework.decorators import action
from django.utils import timezone
from .serializers_admin import (
    InstitutionAdminSerializer,
    UserAdminSerializer,
    CourseAdminReadSerializer,
    CourseAdminWriteSerializer,
    EnrollmentAdminSerializer,
    SessionAdminSerializer,
    DepartmentAdminSerializer,
    SemesterAdminSerializer,
    SectionAdminSerializer
)
from django_filters.rest_framework import DjangoFilterBackend

User = get_user_model()

class AdminInstitutionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = InstitutionAdminSerializer

    def get_permissions(self):
        # Only Global Super Admins can create, edit, or delete institutions
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAuthenticated(), IsGlobalAdmin()]
        return [IsAuthenticated(), IsAdminUser()]

    def get_queryset(self):
        qs = Institution.objects.all().annotate(
            user_count=Count("users", distinct=True),
            course_count=Count("courses", distinct=True)
        )
        # Institution-specific admins are scoped to their assigned institution
        if not self.request.user.is_superuser and self.request.user.institution:
            qs = qs.filter(id=self.request.user.institution.id)
        return qs

class AdminDepartmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = DepartmentAdminSerializer
    
    def get_queryset(self):
        queryset = Department.objects.all()
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(id=self.request.user.department.id)
        
        inst_id = self.request.query_params.get("institution")
        if inst_id:
            queryset = queryset.filter(institution_id=inst_id)
        return queryset

class AdminSemesterViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = SemesterAdminSerializer
    
    def get_queryset(self):
        queryset = Semester.objects.all()
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(department__institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(department=self.request.user.department)
        
        dept_id = self.request.query_params.get("department")
        if dept_id:
            queryset = queryset.filter(department_id=dept_id)
        return queryset

class AdminSectionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = SectionAdminSerializer
    
    def get_queryset(self):
        queryset = Section.objects.all()
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(semester__department__institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(semester__department=self.request.user.department)
        
        sem_id = self.request.query_params.get("semester")
        if sem_id:
            queryset = queryset.filter(semester_id=sem_id)
        return queryset

class AdminUserViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = UserAdminSerializer
    
    def get_queryset(self):
        queryset = User.objects.select_related(
            "institution",
            "department",
            "section",
            "section__semester",
            "section__semester__department",
            "section__semester__department__institution"
        ).all()
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(
                    Q(institution=self.request.user.institution) | 
                    Q(section__semester__department__institution=self.request.user.institution) |
                    Q(department__institution=self.request.user.institution)
                )
            if self.request.user.department:
                queryset = queryset.filter(
                    Q(department=self.request.user.department) |
                    Q(section__semester__department=self.request.user.department)
                )
        
        inst_id = self.request.query_params.get("institution")
        if inst_id:
            queryset = queryset.filter(
                Q(institution_id=inst_id) | 
                Q(section__semester__department__institution_id=inst_id)
            )
        return queryset

    @action(detail=False, methods=["post"], url_path="bulk-generate")
    def bulk_generate(self, request):
        section_id = request.data.get("section_id")
        count = request.data.get("count", 1)
        prefix = request.data.get("prefix", "std_")
        course_id = request.data.get("course_id")
        custom_password = request.data.get("password")
        
        if not section_id:
            return Response({"error": "section_id is required"}, status=status.HTTP_400_BAD_REQUEST)
            
        try:
            section = Section.objects.get(id=section_id)
        except Section.DoesNotExist:
            return Response({"error": "Section not found"}, status=status.HTTP_404_NOT_FOUND)
            
        course = None
        if course_id:
            try:
                course = Course.objects.get(id=course_id)
            except Course.DoesNotExist:
                return Response({"error": "Course not found"}, status=status.HTTP_404_NOT_FOUND)

        # Check permissions for scoped admin
        if not request.user.is_superuser:
            if request.user.institution and section.semester.department.institution != request.user.institution:
                return Response({"error": "Permission denied for this institution"}, status=status.HTTP_403_FORBIDDEN)
            if request.user.department and section.semester.department != request.user.department:
                return Response({"error": "Permission denied for this department"}, status=status.HTTP_403_FORBIDDEN)

        institution = section.semester.department.institution
        domain = institution.slug or "uni"
        
        created_users = []
        import string
        import random
        from django.db import transaction
        
        try:
            with transaction.atomic():
                for i in range(1, int(count) + 1):
                    attempts = 0
                    while attempts < 100:
                        suffix = f"{i:03d}" if int(count) > 1 or attempts > 0 else f"{random.randint(100, 999)}"
                        email = f"{prefix}{suffix}@{domain}.edu".lower()
                        if not User.objects.filter(email=email).exists():
                            break
                        attempts += 1
                        prefix = prefix + str(random.randint(0, 9))
                    
                    if custom_password:
                        password = custom_password
                    else:
                        password_chars = string.ascii_letters + string.digits
                        password = "".join(random.choice(password_chars) for _ in range(8))
                    
                    user = User.objects.create_user(
                        email=email,
                        password=password,
                        role="student",
                        section=section
                    )
                    
                    if course:
                        Enrollment.objects.create(
                            student=user,
                            course=course
                        )

                    created_users.append({
                        "email": email,
                        "password": password,
                        "role": "student",
                        "section_name": section.name,
                        "semester_number": section.semester.number,
                        "department_name": section.semester.department.name,
                        "enrolled_course": course.name if course else None
                    })

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
        return Response({"users": created_users, "message": f"Successfully generated {len(created_users)} students."}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="import-file")
    def import_file(self, request):
        import csv
        import json
        import openpyxl
        from django.db import transaction
        from apps.accounts.views import generate_and_send_otp

        file_obj = request.FILES.get("file")
        dry_run = request.data.get("dry_run", "false").lower() == "true"
        auto_create_structure = request.data.get("auto_create_structure", "false").lower() == "true"
        password_strategy = request.data.get("password_strategy", "auto") # auto, reg_no, custom
        custom_password = request.data.get("custom_password", "")
        course_id = request.data.get("course_id")
        
        column_mapping_str = request.data.get("column_mapping", "{}")
        try:
            column_mapping = json.loads(column_mapping_str)
        except Exception:
            column_mapping = {}

        if not file_obj:
            return Response({"error": "No file uploaded"}, status=status.HTTP_400_BAD_REQUEST)

        filename = file_obj.name.lower()
        rows = []
        headers = []

        try:
            if filename.endswith(".csv"):
                decoded_file = file_obj.read().decode("utf-8").splitlines()
                reader = csv.reader(decoded_file)
                headers = [h.strip() for h in next(reader, [])]
                for r in reader:
                    if any(r):
                        rows.append([cell.strip() for cell in r])
            elif filename.endswith((".xlsx", ".xls")):
                wb = openpyxl.load_workbook(file_obj, read_only=True)
                ws = wb.active
                iter_rows = ws.iter_rows(values_only=True)
                headers = [str(h).strip() if h is not None else "" for h in next(iter_rows, [])]
                for r in iter_rows:
                    if any(r):
                        rows.append([str(cell).strip() if cell is not None else "" for cell in r])
            else:
                return Response({"error": "Unsupported file format. Please upload a CSV or Excel file."}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({"error": f"Failed to parse file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        if not headers:
            return Response({"error": "File is empty or contains no headers"}, status=status.HTTP_400_BAD_REQUEST)

        if request.query_params.get("get_preview") == "true":
            preview_rows = rows[:5]
            return Response({
                "headers": headers,
                "preview_rows": preview_rows,
                "total_rows": len(rows)
            })

        idx_email = headers.index(column_mapping.get("email")) if column_mapping.get("email") in headers else -1
        idx_reg = headers.index(column_mapping.get("registration_number")) if column_mapping.get("registration_number") in headers else -1
        idx_role = headers.index(column_mapping.get("role")) if column_mapping.get("role") in headers else -1
        idx_dept = headers.index(column_mapping.get("department")) if column_mapping.get("department") in headers else -1
        idx_sem = headers.index(column_mapping.get("semester")) if column_mapping.get("semester") in headers else -1
        idx_sec = headers.index(column_mapping.get("section")) if column_mapping.get("section") in headers else -1

        if idx_email == -1:
            return Response({"error": "Email column mapping is required and must match a header in the file."}, status=status.HTTP_400_BAD_REQUEST)

        admin_institution = None if request.user.is_superuser else request.user.institution

        success_list = []
        error_list = []
        
        target_course = None
        if course_id:
            try:
                target_course = Course.objects.get(id=course_id)
                if not request.user.is_superuser and target_course.institution != admin_institution:
                    return Response({"error": "Access denied for target enrollment course."}, status=status.HTTP_403_FORBIDDEN)
            except Course.DoesNotExist:
                return Response({"error": "Selected enrollment course not found."}, status=status.HTTP_404_NOT_FOUND)

        import random
        import string
        from django.core.exceptions import ValidationError
        from django.core.validators import validate_email

        try:
            with transaction.atomic():
                for row_num, row in enumerate(rows, start=2):
                    if len(row) <= idx_email:
                        error_list.append({"row": row_num, "error": "Row is truncated or email cell is missing."})
                        continue

                    email_val = row[idx_email].strip()
                    reg_val = row[idx_reg].strip() if idx_reg != -1 and len(row) > idx_reg else ""
                    role_val = row[idx_role].strip().lower() if idx_role != -1 and len(row) > idx_role else "student"
                    dept_val = row[idx_dept].strip() if idx_dept != -1 and len(row) > idx_dept else ""
                    sem_val = row[idx_sem].strip() if idx_sem != -1 and len(row) > idx_sem else ""
                    sec_val = row[idx_sec].strip() if idx_sec != -1 and len(row) > idx_sec else ""

                    if not role_val or role_val not in ["student", "teacher", "admin"]:
                        role_val = "student"

                    if not email_val:
                        error_list.append({"row": row_num, "error": "Email is empty."})
                        continue
                    try:
                        validate_email(email_val)
                    except ValidationError:
                        error_list.append({"row": row_num, "error": f"Invalid email format: '{email_val}'."})
                        continue

                    row_inst = admin_institution
                    if not row_inst:
                        email_domain = email_val.split("@")[-1].lower()
                        inst_query = Institution.objects.filter(Q(domain__iexact=email_domain) | Q(slug__iexact=email_domain.split(".")[0]))
                        if inst_query.exists():
                            row_inst = inst_query.first()
                        else:
                            error_list.append({"row": row_num, "error": f"Could not determine Institution for email '{email_val}'."})
                            continue

                    if role_val == "student":
                        inst_domain = row_inst.domain
                        if not inst_domain:
                            inst_domain = f"{row_inst.slug}.edu"
                        
                        inst_domain = inst_domain.strip().lower()
                        email_lower = email_val.lower()
                        if not email_lower.endswith(f"@{inst_domain}") and not email_lower.endswith(f".{inst_domain}"):
                            error_list.append({
                                "row": row_num, 
                                "error": f"Student email '{email_val}' must end with institutional domain '@{inst_domain}'."
                            })
                            continue

                    if User.objects.filter(email__iexact=email_val).exists():
                        error_list.append({"row": row_num, "error": f"A user with email '{email_val}' already exists."})
                        continue

                    if reg_val and User.objects.filter(registration_number__iexact=reg_val).exists():
                        error_list.append({"row": row_num, "error": f"Registration number '{reg_val}' already exists."})
                        continue

                    row_dept = None
                    row_sec = None
                    if role_val in ["student", "teacher"] and (dept_val or auto_create_structure):
                        if dept_val:
                            dept_qs = Department.objects.filter(name__iexact=dept_val, institution=row_inst)
                            if dept_qs.exists():
                                row_dept = dept_qs.first()
                            elif auto_create_structure:
                                row_dept = Department.objects.create(name=dept_val, institution=row_inst)
                            else:
                                error_list.append({"row": row_num, "error": f"Department '{dept_val}' does not exist."})
                                continue

                        if role_val == "student" and (sem_val or sec_val or auto_create_structure):
                            row_sem = None
                            if sem_val and row_dept:
                                sem_qs = Semester.objects.filter(number__iexact=sem_val, department=row_dept)
                                if sem_qs.exists():
                                    row_sem = sem_qs.first()
                                elif auto_create_structure:
                                    row_sem = Semester.objects.create(number=sem_val, department=row_dept)
                                else:
                                    error_list.append({"row": row_num, "error": f"Semester '{sem_val}' does not exist in department '{dept_val}'."})
                                    continue

                            if sec_val and row_sem:
                                sec_qs = Section.objects.filter(name__iexact=sec_val, semester=row_sem)
                                if sec_qs.exists():
                                    row_sec = sec_qs.first()
                                elif auto_create_structure:
                                    row_sec = Section.objects.create(name=sec_val, semester=row_sem)
                                else:
                                    error_list.append({"row": row_num, "error": f"Section '{sec_val}' does not exist in Semester '{sem_val}'."})
                                    continue

                    password = ""
                    if password_strategy == "reg_no" and reg_val:
                        password = reg_val
                    elif password_strategy == "custom" and custom_password:
                        password = custom_password
                    else:
                        password_chars = string.ascii_letters + string.digits
                        password = "".join(random.choice(password_chars) for _ in range(8))

                    if not dry_run:
                        user = User.objects.create_user(
                            email=email_val.lower(),
                            password=password,
                            role=role_val,
                            institution=row_inst if role_val != "student" else None,
                            department=row_dept if role_val == "teacher" else None,
                            section=row_sec if role_val == "student" else None,
                            registration_number=reg_val or None,
                            is_email_verified=False
                        )

                        if role_val == "student" and target_course:
                            Enrollment.objects.create(student=user, course=target_course)

                        generate_and_send_otp(user, purpose="verify")

                    success_list.append({
                        "email": email_val,
                        "password": password,
                        "role": role_val,
                        "registration_number": reg_val,
                        "institution": row_inst.name,
                        "department": row_dept.name if row_dept else "",
                        "section": row_sec.name if row_sec else "",
                    })

                if dry_run:
                    raise transaction.TransactionManagementError("Dry run roll back")

        except transaction.TransactionManagementError:
            pass
        except Exception as e:
            return Response({"error": f"An unexpected database error occurred during import: {str(e)}"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({
            "dry_run": dry_run,
            "success_count": len(success_list),
            "error_count": len(error_list),
            "errors": error_list,
            "imported_users": success_list
        }, status=status.HTTP_200_OK)


class AdminCourseViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = CourseAdminReadSerializer

    def get_serializer_class(self):
        if self.action in ["create", "update", "partial_update"]:
            return CourseAdminWriteSerializer
        return CourseAdminReadSerializer

    def get_queryset(self):
        queryset = Course.objects.all().annotate(
            enrollment_count=Count("enrollments", distinct=True)
        )
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(department=self.request.user.department)
        return queryset

class AdminEnrollmentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = EnrollmentAdminSerializer
    
    def get_queryset(self):
        queryset = Enrollment.objects.all()
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(course__institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(course__department=self.request.user.department)
        return queryset

class AdminSessionResetView(views.APIView):
    permission_classes = [IsAuthenticated, IsAdminUser]

    def post(self, request):
        session_id = request.data.get("session_id")
        if not session_id:
            return Response({"error": "session_id is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            session = AttendanceSession.objects.get(id=session_id)
            if not request.user.is_superuser:
                if request.user.institution and session.course.institution != request.user.institution:
                    return Response({"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN)
                if request.user.department and session.course.department != request.user.department:
                    return Response({"error": "Permission denied"}, status=status.HTTP_403_FORBIDDEN)
            
            session.expiry_time = timezone.now()
            session.save()
            session.qr_tokens.all().update(expiry_time=timezone.now())
            return Response({"message": f"Session {session_id} has been reset and expired successfully."})
        except AttendanceSession.DoesNotExist:
            return Response({"error": "Session not found"}, status=status.HTTP_404_NOT_FOUND)

class AdminSessionViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsAdminUser]
    serializer_class = SessionAdminSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["course"]

    def get_queryset(self):
        queryset = AttendanceSession.objects.all().order_by("-start_time")
        if not self.request.user.is_superuser:
            if self.request.user.institution:
                queryset = queryset.filter(course__institution=self.request.user.institution)
            if self.request.user.department:
                queryset = queryset.filter(course__department=self.request.user.department)
        return queryset
